from telebot import types
from models import *
from models import *
from datetime import datetime
from telebot import types
from decorators import admin_required
import tempfile
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl import load_workbook
import tempfile
import os
from models import Category, Subcategory
from working_db import Session
from sqlalchemy import func
from datetime import datetime


def register_admin_handlers(bot):
    
    @bot.message_handler(func=lambda msg: msg.text == "👥 Пользователи")
    @admin_required
    def list_users(message):
        """Список всех пользователей"""
        with Session() as session:
            users = session.query(User).order_by(User.id.desc()).all()
            
            if not users:
                bot.reply_to(message, "📭 Нет пользователей в системе")
                return
        # Создаём временную директорию и файл
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            filepath = tmp.name

        # Создаём Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Пользователи"

        # Заголовки
        headers = ["ID", "Username", "Telegram ID", "Роли"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Заполняем данными
        for user in users:
            roles = []
            if user.is_admin:
                roles.append("Админ")
            if user.is_it_specialist:
                roles.append("ИТ-спец")
            if not roles:
                roles.append("Пользователь")

            ws.append([
                user.id,
                f"@{user.username}" if user.username else "",
                user.telegram_id,
                ", ".join(roles)
            ])

        # Автоширина столбцов (опционально)
        for col in "ABCD":
            ws.column_dimensions[col].width = 20

        # Сохраняем файл
        wb.save(filepath)

        # Отправляем в чат
        with open(filepath, 'rb') as excel_file:
            bot.send_document(
                chat_id=message.chat.id,
                document=excel_file,
                caption="📋 Список всех пользователей (Excel)",
                visible_file_name="Список всех пользователей.xlsx"
            )
        # Удаляем временный файл
        os.unlink(filepath)


    # def start_broadcast(message):
    #     """Начать рассылку"""
    #     msg = bot.reply_to(message, "📢 Введите сообщение для рассылки:")
    #     bot.register_next_step_handler(msg, process_broadcast)

    # def process_broadcast(message):
    #     """Обработать рассылку"""
    #     with Session() as session:
    #         users = session.query(User).all()
    #         success = 0
    #         for user in users:
    #             try:
    #                 bot.send_message(user.telegram_id, f"📢 <b>Рассылка</b>:\n\n{message.text}", parse_mode='HTML')
    #                 success += 1
    #             except:
    #                 continue
            
    #         bot.reply_to(message, f"✅ Рассылка завершена\nОтправлено: {success}/{len(users)}")
    
    # @bot.message_handler(commands=['delete_ticket'])
    # def handle_delete_ticket(message):
    #     """Обработчик удаления заявки с подтверждением"""
    #     try:
    #         args = message.text.split()
    #         if len(args) < 2:
    #             bot.reply_to(message, "❌ Формат: /delete_ticket <ID заявки>")
    #             return

    #         ticket_id = int(args[1])
            
    #         with Session() as session:

    #             ticket = session.query(Ticket).filter_by(id=ticket_id).first()
                
    #             if not ticket:
    #                 bot.reply_to(message, f"❌ Заявка #{ticket_id} не найдена")
    #                 return

    #             # Создаем клавиатуру подтверждения
    #             markup = types.InlineKeyboardMarkup()
    #             markup.add(
    #                 types.InlineKeyboardButton("✅ Да, удалить", callback_data=f"confirm_del_{ticket_id}"),
    #                 types.InlineKeyboardButton("❌ Отмена", callback_data="cancel_del")
    #             )
                
    #             bot.send_message(
    #                 message.chat.id,
    #                 f"⚠️ Вы уверены, что хотите полностью удалить заявку #{ticket_id}?\n"
    #                 f"Категория: {ticket.subcategory.category.name}\n"
    #                 f"Пользователь: @{ticket.user.username}\n"
    #                 f"Это действие нельзя отменить!",
    #                 reply_markup=markup
    #             )

    #     except ValueError:
    #         bot.reply_to(message, "❌ Некорректный ID заявки")
    #     except Exception as e:
    #         bot.reply_to(message, f"⚠️ Ошибка: {str(e)}")

    # @bot.callback_query_handler(func=lambda call: call.data.startswith('confirm_del_'))
    # def confirm_delete(call):
    #     """Подтверждение удаления"""
    #     ticket_id = int(call.data.split('_')[-1])
        
    #     with Session() as session:
    #         try:
    #             ticket = session.query(Ticket).filter_by(id=ticket_id).first()
    #             if ticket:
    #                 # Удаляем файлы, если они есть
    #                 if ticket.file_path and os.path.exists(ticket.file_path):
    #                     os.remove(ticket.file_path)
                    
    #                 session.delete(ticket)
    #                 session.commit()
                    
    #                 bot.edit_message_text(
    #                     f"✅ Заявка #{ticket_id} полностью удалена",
    #                     call.message.chat.id,
    #                     call.message.message_id
    #                 )
    #             else:
    #                 bot.answer_callback_query(call.id, "Заявка уже удалена")

    #         except Exception as e:
    #             session.rollback()
    #             bot.edit_message_text(
    #                 f"❌ Ошибка удаления: {str(e)}",
    #                 call.message.chat.id,
    #                 call.message.message_id
                # )
    @bot.message_handler(func=lambda msg: msg.text == "📋 Загрузить категории")
    @admin_required
    def handle_upload_categories(message):
        """Запрашивает Excel-файл с категориями от админа"""
        msg = bot.reply_to(
            message,
            "📤 Пожалуйста, отправьте Excel-файл (.xlsx) с категориями.")
        bot.register_next_step_handler(msg, _process_uploaded_excel)
        
    def _process_uploaded_excel(message):
        """Полная синхронизация категорий из Excel-файла 'категории.xlsx'"""
        if message.content_type != 'document':
            bot.reply_to(message, "❌ Пожалуйста, отправьте файл Excel (.xlsx)")
            return

        filename = message.document.file_name
        if filename != "категории.xlsx":
            bot.reply_to(
                message,
                "❌ Неверное имя файла!\n"
                "Файл должен называться точно: <b>категории.xlsx</b>",
                parse_mode='HTML'
            )
            return

        try:
            # Скачиваем файл
            file_info = bot.get_file(message.document.file_id)
            downloaded_file = bot.download_file(file_info.file_path)

            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                tmp.write(downloaded_file)
                tmp_path = tmp.name

            wb = load_workbook(tmp_path)
            os.unlink(tmp_path)

            # Собираем данные из Excel
            excel_data = {}
            for sheet_name in wb.sheetnames:
                category_name = sheet_name.strip()
                if not category_name:
                    continue
                excel_data[category_name] = []
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                    subcat_name = row[0]
                    recommendation = row[1] if row[1] is not None else ""
                    if subcat_name:
                        excel_data[category_name].append({
                            'name': str(subcat_name).strip(),
                            'recommendation': str(recommendation).strip()
                        })

            # Работаем с БД
            with Session() as session:
                # === ШАГ 1: Обновляем/добавляем категории и подкатегории ===
                db_categories = {cat.name: cat for cat in session.query(Category).all()}
                added_categories = 0
                added_subcategories = 0
                updated_recommendations = 0

                for cat_name, subcats in excel_data.items():
                    # Категория
                    category = db_categories.get(cat_name)
                    if not category:
                        category = Category(name=cat_name)
                        session.add(category)
                        session.flush()
                        added_categories += 1
                    db_categories[cat_name] = category  # обновляем после flush

                    # Подкатегории
                    db_subcats = {
                        sub.name: sub
                        for sub in session.query(Subcategory).filter_by(category_id=category.id).all()
                    }

                    for sub_data in subcats:
                        sub_name = sub_data['name']
                        rec = sub_data['recommendation']

                        sub = db_subcats.get(sub_name)
                        if not sub:
                            sub = Subcategory(
                                category_id=category.id,
                                name=sub_name,
                                recommendation=rec
                            )
                            session.add(sub)
                            added_subcategories += 1
                        else:
                            if sub.recommendation != rec:
                                sub.recommendation = rec
                                updated_recommendations += 1

                session.flush()  # Чтобы все ID были актуальны

                # === ШАГ 2: Удаляем ЛИШНИЕ подкатегории (только если нет заявок) ===
                deleted_subcategories = 0
                for cat_name, category in db_categories.items():
                    if cat_name not in excel_data:
                        continue  # Эта категория будет удалена позже (если возможно)
                    db_subcats_in_cat = {
                        sub.name: sub
                        for sub in session.query(Subcategory).filter_by(category_id=category.id).all()
                    }
                    excel_sub_names = {sub['name'] for sub in excel_data[cat_name]}

                    for sub_name, sub in db_subcats_in_cat.items():
                        if sub_name not in excel_sub_names:
                            # Проверяем, есть ли заявки
                            ticket_exists = session.query(Ticket).filter_by(subcategory_id=sub.id).first()
                            if ticket_exists:
                                print(f"⚠️ Не удаляем подкатегорию '{sub_name}' — есть заявки")
                            else:
                                session.delete(sub)
                                deleted_subcategories += 1

                # === ШАГ 3: Удаляем ЛИШНИЕ категории (только если нет подкатегорий) ===
                deleted_categories = 0
                all_db_categories = session.query(Category).all()
                excel_cat_names = set(excel_data.keys())

                for category in all_db_categories:
                    if category.name not in excel_cat_names:
                        # Проверяем, остались ли подкатегории
                        subcat_count = session.query(Subcategory).filter_by(category_id=category.id).count()
                        if subcat_count == 0:
                            session.delete(category)
                            deleted_categories += 1
                        else:
                            print(f"⚠️ Не удаляем категорию '{category.name}' — остались подкатегории")

                session.commit()

            bot.reply_to(
                message,
                f"✅ Синхронизация завершена!\n\n"
                f"🆕 Добавлено категорий: {added_categories}\n"
                f"🆕 Добавлено подкатегорий: {added_subcategories}\n"
                f"🔄 Обновлено рекомендаций: {updated_recommendations}\n"
                f"🗑️ Удалено подкатегорий: {deleted_subcategories}\n"
                f"🗑️ Удалено категорий: {deleted_categories}\n\n"
                f"⚠️ Элементы с заявками не удалялись.",
                parse_mode='HTML'
            )

        except Exception as e:
            error_msg = f"❌ Ошибка синхронизации: {str(e)}"
            bot.reply_to(message, error_msg)
            print("Ошибка синхронизации категорий:", e)
            if 'tmp_path' in locals() and os.path.exists(tmp_path):
                os.unlink(tmp_path)                

    @bot.message_handler(func=lambda msg: msg.text == "Скачать файл с категориями")
    @admin_required
    def download_categories_excel(message):
        """Генерирует и отправляет Excel-файл со всеми категориями и подкатегориями"""
        try:
            with Session() as session:
                categories = session.query(Category).order_by(Category.name).all()

                if not categories:
                    bot.reply_to(message, "📭 В базе нет категорий.")
                    return

                # Создаём новую книгу Excel
                wb = Workbook()
                # Удаляем дефолтный лист
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])

                for category in categories:
                    # Ограничиваем название листа до 31 символа (ограничение Excel)
                    sheet_title = category.name[:31]
                    ws = wb.create_sheet(title=sheet_title)

                    # Заголовки
                    ws["A1"] = "Подкатегория"
                    ws["B1"] = "Рекомендация"
                    for cell in ws[1]:
                        cell.font = Font(bold=True)

                    # Заполняем подкатегории
                    row = 2
                    for subcat in sorted(category.subcategories, key=lambda x: x.name):
                        ws[f"A{row}"] = subcat.name
                        ws[f"B{row}"] = subcat.recommendation
                        row += 1

                    # Автоширина столбцов (ограничим до 100 символов)
                    for col_letter in ["A", "B"]:
                        max_length = 0
                        for cell in ws[col_letter]:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        adjusted_width = min(max_length + 2, 100)
                        ws.column_dimensions[col_letter].width = adjusted_width

                # Сохраняем во временный файл
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    wb.save(tmp.name)
                    tmp_path = tmp.name

            # Отправляем файл
            with open(tmp_path, 'rb') as excel_file:
                bot.send_document(
                    chat_id=message.chat.id,
                    document=excel_file,
                    caption="📁 Текущие категории и рекомендации",
                    visible_file_name="категории.xlsx"
                )

            # Удаляем временный файл
            os.unlink(tmp_path)

        except Exception as e:
            bot.reply_to(message, f"❌ Ошибка генерации файла: {str(e)}")
            print("Ошибка при создании Excel:", e)
            if 'tmp_path' in locals() and os.path.exists(tmp_path):
                os.unlink(tmp_path)                


    @bot.message_handler(func=lambda msg: msg.text == "🛠 Управление ИТ-специалистами")
    @admin_required
    def manage_it_specialists_menu(message):
        """Показывает меню: добавить или удалить IT-спеца"""
        markup = types.InlineKeyboardMarkup()
        markup.row(
            types.InlineKeyboardButton("➕ Добавить", callback_data="it_add"),
            types.InlineKeyboardButton("➖ Удалить", callback_data="it_remove")
        )
        bot.send_message(
            message.chat.id,
            "👨‍💻 Выберите действие:",
            reply_markup=markup
        )


    # === ДОБАВЛЕНИЕ ===

    @bot.callback_query_handler(func=lambda call: call.data == "it_add")
    @admin_required
    def ask_add_it_specialist(call):
        bot.answer_callback_query(call.id)
        msg = bot.send_message(
            call.message.chat.id,
            "Введите <b>Telegram ID</b> или <b>username</b> (без @) пользователя, который уже запускал бота:",
            parse_mode='HTML'
        )
        bot.register_next_step_handler(msg, process_add_it_input)


    def process_add_it_input(message):
        text = message.text.strip()
        if not text:
            bot.reply_to(message, "❌ Ввод пуст. Используйте меню снова.")
            return

        try:
            with Session() as session:
                # Ищем пользователя
                user = None
                if text.isdigit():
                    user = session.query(User).filter_by(telegram_id=int(text)).first()
                else:
                    user = session.query(User).filter_by(username=text).first()

                if not user:
                    bot.reply_to(
                        message,
                        "❌ Пользователь не найден.\n"
                        "Он должен был написать боту /start хотя бы один раз."
                    )
                    return

                if user.is_it_specialist:
                    name = f"@{user.username}" if user.username else f"ID {user.telegram_id}"
                    bot.reply_to(message, f"ℹ️ {name} уже является IT-специалистом.")
                    return

                # Подтверждение
                display = f"@{user.username}" if user.username else f"ID {user.telegram_id}"
                markup = types.InlineKeyboardMarkup()
                markup.row(
                    types.InlineKeyboardButton("✅ Да", callback_data=f"confirm_add_{user.telegram_id}"),
                    types.InlineKeyboardButton("❌ Нет", callback_data="cancel_it")
                )
                bot.send_message(
                    message.chat.id,
                    f"Назначить {display} IT-специалистом?",
                    reply_markup=markup
                )

        except Exception as e:
            bot.reply_to(message, f"❌ Ошибка: {str(e)}")


    @bot.callback_query_handler(func=lambda call: call.data.startswith("confirm_add_"))
    @admin_required
    def confirm_add_it(call):
        telegram_id = int(call.data.split("_")[-1])
        with Session() as session:
            user = session.query(User).filter_by(telegram_id=telegram_id).first()
            if not user or user.is_it_specialist:
                bot.edit_message_text("❌ Невозможно назначить.", call.message.chat.id, call.message.message_id)
                return

            user.is_it_specialist = True
            session.commit()

            name = f"@{user.username}" if user.username else f"ID {user.telegram_id}"
            bot.edit_message_text(
                f"✅ {name} теперь IT-специалист!",
                call.message.chat.id,
                call.message.message_id
            )

            # Уведомление (опционально)
            try:
                bot.send_message(user.telegram_id, "👨‍💻 Вам выдана роль IT-специалиста! Ознакомтись с документацией (тут можно положить ссылку)")
            except Exception as e:
                print('Общибка в отправке IT-специалисту в назначении прав {e}')


    # === УДАЛЕНИЕ ===
    @bot.callback_query_handler(func=lambda call: call.data == "it_remove")
    @admin_required
    def ask_remove_it_specialist(call):
        bot.answer_callback_query(call.id)
        msg = bot.send_message(
            call.message.chat.id,
            "Введите <b>Telegram ID</b> или <b>username</b> (без @) IT-специалиста для удаления:",
            parse_mode='HTML'
        )
        bot.register_next_step_handler(msg, process_remove_it_input)


    def process_remove_it_input(message):
        text = message.text.strip()
        if not text:
            bot.reply_to(message, "❌ Ввод пуст.")
            return

        try:
            with Session() as session:
                user = None
                if text.isdigit():
                    user = session.query(User).filter_by(telegram_id=int(text)).first()
                else:
                    user = session.query(User).filter_by(username=text).first()

                if not user:
                    bot.reply_to(message, "❌ Пользователь не найден.")
                    return

                if not user.is_it_specialist:
                    name = f"@{user.username}" if user.username else f"ID {user.telegram_id}"
                    bot.reply_to(message, f"❌ {name} не является IT-специалистом.")
                    return

                display = f"@{user.username}" if user.username else f"ID {user.telegram_id}"
                markup = types.InlineKeyboardMarkup()
                markup.row(
                    types.InlineKeyboardButton("✅ Да", callback_data=f"confirm_remove_{user.telegram_id}"),
                    types.InlineKeyboardButton("❌ Нет", callback_data="cancel_it")
                )
                bot.send_message(
                    message.chat.id,
                    f"Лишить {display} прав IT-специалиста?",
                    reply_markup=markup
                )

        except Exception as e:
            bot.reply_to(message, f"❌ Ошибка: {str(e)}")


    @bot.callback_query_handler(func=lambda call: call.data.startswith("confirm_remove_"))
    @admin_required
    def confirm_remove_it(call):
        telegram_id = int(call.data.split("_")[-1])
        with Session() as session:
            user = session.query(User).filter_by(telegram_id=telegram_id).first()
            if not user or not user.is_it_specialist:
                bot.edit_message_text("❌ Пользователь не является IT-специалистом.", call.message.chat.id, call.message.message_id)
                return

            user.is_it_specialist = False
            session.commit()

            name = f"@{user.username}" if user.username else f"ID {user.telegram_id}"
            bot.edit_message_text(
                f"✅ {name} больше не IT-специалист.",
                call.message.chat.id,
                call.message.message_id
            )

            try:
                bot.send_message(user.telegram_id, "👨‍💻 У вас отозвана роль IT-специалиста.")
            except:
                pass


    # Общий обработчик отмены
    @bot.callback_query_handler(func=lambda call: call.data == "cancel_it")
    def cancel_it_action(call):
        bot.edit_message_text("❌ Отменено.", call.message.chat.id, call.message.message_id)
            
    @bot.message_handler(func=lambda msg: msg.text == "📊 Статистика")
    @admin_required
    def show_stats_excel(message):
        """Генерирует Excel-файл со статистикой и диаграммами"""
        try:
            with Session() as session:
                # === Основные данные ===
                total_tickets = session.query(Ticket).count()
                if total_tickets == 0:
                    bot.reply_to(message, "📭 Нет заявок для анализа.")
                    return

                # Статусы
                status_counts = dict(session.query(Ticket.status, func.count(Ticket.id)).group_by(Ticket.status).all())

                # По категориям
                cat_data = session.query(
                    Category.name,
                    func.count(Ticket.id)
                ).join(Subcategory).join(Ticket).group_by(Category.id).all()

                # По подкатегориям (ТОП-20)
                subcat_data = session.query(
                    Subcategory.name,
                    func.count(Ticket.id)
                ).join(Ticket).group_by(Subcategory.id).order_by(func.count(Ticket.id).desc()).limit(20).all()

                # Динамика по месяцам
                monthly_data = session.query(
                    func.strftime('%Y-%m', Ticket.created_at).label('month'),
                    func.count(Ticket.id)
                ).group_by('month').order_by('month').all()

                # Дата первой заявки
                first_ticket = session.query(func.min(Ticket.created_at)).scalar()

            # === Создаём Excel ===
            wb = Workbook()
            ws_summary = wb.active
            ws_summary.title = "Сводка"

            # --- Общая сводка ---
            ws_summary["A1"] = "📊 Статистика IT-заявок"
            ws_summary["A1"].font = Font(size=16, bold=True)
            ws_summary["A3"] = "Всего заявок:"
            ws_summary["B3"] = total_tickets
            ws_summary["A4"] = "В работе:"
            ws_summary["B4"] = status_counts.get("В работе", 0)
            ws_summary["A5"] = "Открыто:"
            ws_summary["B5"] = status_counts.get("Открыт", 0)
            ws_summary["A6"] = "Закрыто:"
            ws_summary["B6"] = status_counts.get("Закрыт", 0)
            ws_summary["A7"] = "Период:"
            ws_summary["B7"] = f"{first_ticket.strftime('%d.%m.%Y')} — {datetime.now().strftime('%d.%m.%Y')}"

            for row in ws_summary['A3:B7']:
                row[0].font = Font(bold=True)

            # --- Лист: Категории ---
            ws_cats = wb.create_sheet("По категориям")
            ws_cats["A1"] = "Категория"
            ws_cats["B1"] = "Заявок"
            for cell in ws_cats[1]:
                cell.font = Font(bold=True)

            for i, (name, count) in enumerate(cat_data, start=2):
                ws_cats[f"A{i}"] = name
                ws_cats[f"B{i}"] = count

            # Диаграмма по категориям
            from openpyxl.chart import PieChart, Reference
            if len(cat_data) > 0:
                pie = PieChart()
                labels = Reference(ws_cats, min_col=1, min_row=2, max_row=len(cat_data)+1)
                data = Reference(ws_cats, min_col=2, min_row=1, max_row=len(cat_data)+1)
                pie.add_data(data, titles_from_data=True)
                pie.set_categories(labels)
                pie.title = "Заявки по категориям"
                ws_cats.add_chart(pie, "D2")

            # --- Лист: Подкатегории ---
            ws_subcats = wb.create_sheet("По подкатегориям")
            ws_subcats["A1"] = "Подкатегория"
            ws_subcats["B1"] = "Заявок"
            for cell in ws_subcats[1]:
                cell.font = Font(bold=True)

            for i, (name, count) in enumerate(subcat_data, start=2):
                ws_subcats[f"A{i}"] = name
                ws_subcats[f"B{i}"] = count

            # Гистограмма (ТОП-10)
            if len(subcat_data) > 0:
                from openpyxl.chart import BarChart
                bar = BarChart()
                bar.type = "col"
                bar.style = 10
                bar.title = "ТОП-10 подкатегорий"
                bar.y_axis.title = 'Количество'
                bar.x_axis.title = 'Подкатегория'

                data = Reference(ws_subcats, min_col=2, min_row=1, max_row=min(11, len(subcat_data)+1))
                cats = Reference(ws_subcats, min_col=1, min_row=2, max_row=min(11, len(subcat_data)+1))
                bar.add_data(data, titles_from_data=True)
                bar.set_categories(cats)
                ws_subcats.add_chart(bar, "D2")

            # --- Лист: Статусы ---
            ws_status = wb.create_sheet("Статусы")
            ws_status["A1"] = "Статус"
            ws_status["B1"] = "Количество"
            for cell in ws_status[1]:
                cell.font = Font(bold=True)

            statuses = ["Открыт", "В работе", "Закрыт", "Ожидает уточнений"]
            for i, status in enumerate(statuses, start=2):
                ws_status[f"A{i}"] = status
                ws_status[f"B{i}"] = status_counts.get(status, 0)

            # Круговая диаграмма статусов
            pie2 = PieChart()
            labels2 = Reference(ws_status, min_col=1, min_row=2, max_row=5)
            data2 = Reference(ws_status, min_col=2, min_row=1, max_row=5)
            pie2.add_data(data2, titles_from_data=True)
            pie2.set_categories(labels2)
            pie2.title = "Распределение по статусам"
            ws_status.add_chart(pie2, "D2")

            # --- Лист: Динамика ---
            ws_trend = wb.create_sheet("Динамика")
            ws_trend["A1"] = "Месяц"
            ws_trend["B1"] = "Заявок"
            for cell in ws_trend[1]:
                cell.font = Font(bold=True)

            for i, (month, count) in enumerate(monthly_data, start=2):
                ws_trend[f"A{i}"] = month
                ws_trend[f"B{i}"] = count

            # Линейный график
            if len(monthly_data) > 1:
                from openpyxl.chart import LineChart
                line = LineChart()
                line.title = "Динамика заявок по месяцам"
                line.style = 13
                line.y_axis.title = "Количество"
                line.x_axis.title = "Месяц"

                data = Reference(ws_trend, min_col=2, min_row=1, max_row=len(monthly_data)+1)
                cats = Reference(ws_trend, min_col=1, min_row=2, max_row=len(monthly_data)+1)
                line.add_data(data, titles_from_data=True)
                line.set_categories(cats)
                ws_trend.add_chart(line, "D2")

            # === Сохраняем и отправляем ===
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                wb.save(tmp.name)
                tmp_path = tmp.name

            with open(tmp_path, 'rb') as f:
                bot.send_document(
                    message.chat.id,
                    f,
                    caption="📊 Полная статистика заявок (Excel с диаграммами)",
                    visible_file_name="Статистика_заявок.xlsx"
                )
            os.unlink(tmp_path)

        except Exception as e:
            bot.reply_to(message, f"❌ Ошибка генерации отчёта: {str(e)}")
            print("Ошибка статистики:", e)
            if 'tmp_path' in locals():
                try:
                    os.unlink(tmp_path)
                except:
                    pass